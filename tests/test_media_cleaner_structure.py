import io
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from process import media_cleaner
from process.media_cleaner.settings import render_ollama_user_prompt
from process.template_schema import TEMPLATE_COLUMNS


class MediaCleanerStructureTests(unittest.TestCase):
    def test_text_schema_is_loaded_through_public_package(self):
        self.assertEqual(media_cleaner.TARGET_COLUMNS, TEMPLATE_COLUMNS)
        self.assertEqual(media_cleaner.TARGET_COLUMNS[0], "Region")
        self.assertIn("GA Session Qty", media_cleaner.TARGET_COLUMNS)
        self.assertEqual(
            set(media_cleaner.TARGET_COLUMNS),
            set(media_cleaner.TARGET_DESCRIPTIONS),
        )

    def test_editable_prompt_placeholders_render(self):
        prompt = render_ollama_user_prompt(
            target_descriptions='{"Date": "reporting date"}',
            options='[{"row": 1}]',
        )
        self.assertIn("reporting date", prompt)
        self.assertIn('"row": 1', prompt)

    def test_selected_sheets_are_cleaned_and_audited_independently(self):
        workbook = Workbook()
        first = workbook.active
        first.title = "Meta"
        first.append(["日期", "廣告活動名稱", "曝光次數"])
        first.append(["2026-08-01", "Campaign A", 100])
        second = workbook.create_sheet("Google")
        second.append(["Date", "Campaign name", "Impressions"])
        second.append(["2026-08-02", "Campaign B", 200])

        memory_file = io.BytesIO()
        workbook.save(memory_file)
        workbook.close()
        workbook_bytes = memory_file.getvalue()

        self.assertEqual(
            media_cleaner.list_workbook_sheets(workbook_bytes),
            ["Meta", "Google"],
        )

        with tempfile.TemporaryDirectory() as temp_dir:
            input_path = Path(temp_dir) / "vendor_report.xlsx"
            output_path = Path(temp_dir) / "vendor_report_cleaned.xlsx"
            input_path.write_bytes(workbook_bytes)

            audit, row_count = media_cleaner.clean_workbook_sheets(
                input_path=input_path,
                output_path=output_path,
                aliases=media_cleaner.read_dictionary(None),
                scan_rows=20,
                sheet_names=["Google", "Meta"],
                ollama=media_cleaner.OllamaConfig(enabled=False),
            )

            self.assertEqual(row_count, 2)
            self.assertEqual([record.sheet for record in audit], ["Google", "Meta"])
            self.assertTrue(all(record.status == "成功" for record in audit))

            cleaned = load_workbook(output_path, data_only=True, read_only=True)
            try:
                rows = list(cleaned["cleaned_data"].iter_rows(values_only=True))
            finally:
                cleaned.close()
            self.assertEqual(
                list(rows[0]),
                ["Date", "Campaign name", "Impressions"],
            )
            campaign_index = rows[0].index("Campaign name")
            self.assertEqual(
                [row[campaign_index] for row in rows[1:]],
                ["Campaign B", "Campaign A"],
            )

    def test_non_impression_report_maps_all_supported_template_fields(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "GA4"
        worksheet.append(
            ["日期", "使用者", "新使用者", "工作階段", "跳出率", "備註"]
        )
        worksheet.append(["2026-08-01", 120, 80, 150, 0.25, "keep in audit only"])
        memory_file = io.BytesIO()
        workbook.save(memory_file)
        workbook.close()

        with tempfile.TemporaryDirectory() as temp_dir:
            input_path = Path(temp_dir) / "ga4.xlsx"
            output_path = Path(temp_dir) / "ga4_cleaned.xlsx"
            input_path.write_bytes(memory_file.getvalue())

            audit, row_count = media_cleaner.clean_workbook_sheets(
                input_path=input_path,
                output_path=output_path,
                aliases=media_cleaner.read_dictionary(None),
                scan_rows=20,
                sheet_names=["GA4"],
                ollama=media_cleaner.OllamaConfig(enabled=False),
            )

            self.assertEqual(row_count, 1)
            self.assertEqual(audit[0].status, "成功")
            self.assertIn("備註", audit[0].unmapped_columns)
            cleaned = load_workbook(output_path, data_only=True, read_only=True)
            try:
                headers = [cell.value for cell in cleaned["cleaned_data"][1]]
            finally:
                cleaned.close()
            self.assertEqual(
                headers,
                ["Date", "User number", "New user", "Working session", "Bounce Rate"],
            )
            self.assertTrue(set(headers).issubset(TEMPLATE_COLUMNS))

    def test_horizontal_total_rows_are_excluded(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Daily"
        worksheet.append(["Date", "Campaign name", "Impressions"])
        worksheet.append(["2026-08-01", "Campaign A", 100])
        worksheet.append(["Grand Total", None, 100])
        worksheet.append(["2026-08-02", "Total", 100])
        worksheet.append([None, "Subtotal", 100])
        worksheet.append([None, "總計：", 100])
        worksheet.append([None, "小計", 100])
        worksheet.append(["2026-08-03", "Total Awareness Campaign", 200])
        memory_file = io.BytesIO()
        workbook.save(memory_file)
        workbook.close()

        with tempfile.TemporaryDirectory() as temp_dir:
            input_path = Path(temp_dir) / "report_with_totals.xlsx"
            output_path = Path(temp_dir) / "report_with_totals_cleaned.xlsx"
            input_path.write_bytes(memory_file.getvalue())

            audit, row_count = media_cleaner.clean_workbook_sheets(
                input_path=input_path,
                output_path=output_path,
                aliases=media_cleaner.read_dictionary(None),
                scan_rows=20,
                sheet_names=["Daily"],
                ollama=media_cleaner.OllamaConfig(enabled=False),
            )

            self.assertEqual(row_count, 2)
            self.assertEqual(audit[0].status, "成功")
            cleaned = load_workbook(output_path, data_only=True, read_only=True)
            try:
                rows = list(cleaned["cleaned_data"].iter_rows(values_only=True))
            finally:
                cleaned.close()
            campaign_index = rows[0].index("Campaign name")
            self.assertEqual(
                [row[campaign_index] for row in rows[1:]],
                ["Campaign A", "Total Awareness Campaign"],
            )

    def test_every_template_column_can_be_mapped_and_output(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "All Fields"
        worksheet.append(TEMPLATE_COLUMNS)
        values = [1] * len(TEMPLATE_COLUMNS)
        values[TEMPLATE_COLUMNS.index("Date")] = "2026-08-01"
        worksheet.append(values)
        memory_file = io.BytesIO()
        workbook.save(memory_file)
        workbook.close()

        with tempfile.TemporaryDirectory() as temp_dir:
            input_path = Path(temp_dir) / "all_fields.xlsx"
            output_path = Path(temp_dir) / "all_fields_cleaned.xlsx"
            input_path.write_bytes(memory_file.getvalue())
            audit, row_count = media_cleaner.clean_workbook_sheets(
                input_path=input_path,
                output_path=output_path,
                aliases=media_cleaner.read_dictionary(None),
                scan_rows=5,
                sheet_names=["All Fields"],
                ollama=media_cleaner.OllamaConfig(enabled=False),
            )

            self.assertEqual(row_count, 1)
            self.assertEqual(len(audit[0].mapped_columns), len(TEMPLATE_COLUMNS))
            cleaned = load_workbook(output_path, data_only=True, read_only=True)
            try:
                headers = [cell.value for cell in cleaned["cleaned_data"][1]]
            finally:
                cleaned.close()
            self.assertEqual(headers, TEMPLATE_COLUMNS)

    def test_consolidated_workbook_contains_data_and_audit_sheets(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_root = Path(temp_dir)
            cleaned_paths = []
            for index, media in enumerate(("Meta", "Google"), start=1):
                workbook = Workbook()
                worksheet = workbook.active
                worksheet.title = "cleaned_data"
                metric = "Impressions" if index == 1 else "Clicks (all)"
                worksheet.append(["Date", "Media", metric])
                worksheet.append([f"2026-08-0{index}", media, index * 100])
                cleaned_path = temp_root / f"{media.lower()}_cleaned.xlsx"
                workbook.save(cleaned_path)
                workbook.close()
                cleaned_paths.append(cleaned_path)

            records = [
                media_cleaner.AuditRecord(
                    input_file="meta.xlsx",
                    sheet="Meta",
                    header_row=1,
                    status="成功",
                    score=9.5,
                    data_rows=1,
                    mapped_columns={"日期": "Date", "曝光次數": "Impressions"},
                    unmapped_columns=["備註"],
                    output_file="meta_cleaned.xlsx",
                ),
                media_cleaner.AuditRecord(
                    input_file="google.xlsx",
                    sheet="Google",
                    header_row=1,
                    status="成功",
                    score=9.0,
                    data_rows=1,
                    mapped_columns={"Date": "Date", "Clicks": "Clicks (all)"},
                    unmapped_columns=[],
                    output_file="google_cleaned.xlsx",
                ),
            ]
            output_path = temp_root / "cleaned_media_results.xlsx"

            row_count = media_cleaner.consolidate_cleaned_workbooks(
                cleaned_paths=cleaned_paths,
                records=records,
                output_path=output_path,
            )

            self.assertEqual(row_count, 2)
            result = load_workbook(output_path, data_only=True, read_only=True)
            try:
                self.assertEqual(
                    result.sheetnames,
                    ["cleaned_data", "cleaning_audit", "unmapped_columns"],
                )
                cleaned_rows = list(
                    result["cleaned_data"].iter_rows(values_only=True)
                )
                self.assertEqual(
                    list(cleaned_rows[0]),
                    ["Date", "Media", "Impressions", "Clicks (all)"],
                )
                media_index = cleaned_rows[0].index("Media")
                self.assertEqual(
                    [row[media_index] for row in cleaned_rows[1:]],
                    ["Meta", "Google"],
                )
                self.assertEqual(result["cleaning_audit"].max_row, 3)
                unmapped_rows = list(
                    result["unmapped_columns"].iter_rows(values_only=True)
                )
                self.assertEqual(unmapped_rows[1][-1], "備註")
            finally:
                result.close()


if __name__ == "__main__":
    unittest.main()
