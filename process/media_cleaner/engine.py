#!/usr/bin/env python3
"""Detect and standardize media-report tables in otherwise unknown Excel files.

The detector asks a local Ollama model to identify and map the header row,
validates the rows below it, then exports the detected table using the system
schema. If Ollama is unavailable, the original deterministic rules are used.

Examples:
    python -m process.media_cleaner input.xlsx
    python -m process.media_cleaner "0_其他資料/Raw Data_for Louis" -o cleaned_output
"""

from __future__ import annotations

import argparse
import csv
import io
import json
import math
import re
import sys
import urllib.error
import urllib.request
from dataclasses import asdict, dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from .settings import (
    ALIASES,
    OLLAMA_SYSTEM_PROMPT,
    TARGET_COLUMNS,
    TARGET_DESCRIPTIONS,
    render_ollama_user_prompt,
)

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError as exc:  # pragma: no cover - gives a useful CLI error
    raise SystemExit("缺少 openpyxl；請先執行：pip install openpyxl") from exc


SUMMARY_MARKERS = {
    "total",
    "grand total",
    "overall total",
    "subtotal",
    "sub total",
    "合計",
    "總計",
    "小計",
}
PREFERRED_SHEETS = ("daily", "每日", "by day", "byday")

# ---------------------------------------------------------------------------
# Internal data structures
# ---------------------------------------------------------------------------

@dataclass
class Candidate:
    """A row that looks like the header of a real data table."""

    sheet: str
    header_row: int
    score: float
    headers: list[str]
    mapped: dict[int, str]
    unmapped: list[str]
    data_rows: int


@dataclass
class AuditRecord:
    """Information written to cleaning_audit.json after each attempted import."""

    input_file: str
    sheet: str
    header_row: int | None
    status: str
    score: float | None
    data_rows: int
    mapped_columns: dict[str, str]
    unmapped_columns: list[str]
    output_file: str | None


@dataclass
class OllamaConfig:
    """Connection settings for local LLM-assisted header detection."""

    model: str = "qwen3.5:9b"
    url: str = "http://127.0.0.1:11434"
    timeout: float = 120.0
    enabled: bool = True
    warned: bool = False


# ---------------------------------------------------------------------------
# Small value/type helpers
# ---------------------------------------------------------------------------

def clean_text(value: Any) -> str:
    """Convert a cell to trimmed single-line text for comparison/reporting."""
    if value is None:
        return ""
    text = str(value).replace("\n", " ").replace("\r", " ").strip()
    return re.sub(r"\s+", " ", text)


def normalize(value: Any) -> str:
    """Normalize spelling/spacing so visually equivalent headers will match."""
    text = clean_text(value).lower()
    text = text.replace("（", "(").replace("）", ")").replace("％", "%")
    return re.sub(r"[\s_:/：,，。·・\-]+", "", text)


def is_blank(value: Any) -> bool:
    """Treat None, empty strings and vendor dash placeholders as blank."""
    return (
        value is None
        or (isinstance(value, float) and math.isnan(value))
        or clean_text(value) in {"", "--", "—"}
    )


def is_error(value: Any) -> bool:
    """Return True for Excel error strings such as #DIV/0!."""
    return isinstance(value, str) and value.startswith("#")


def is_summary_row(values: Iterable[Any]) -> bool:
    """Return True when any cell explicitly labels a horizontal total row."""
    for value in values:
        if not isinstance(value, str):
            continue
        label = clean_text(value).casefold()
        label = re.sub(r"^[\s:：\-–—_]+|[\s:：\-–—_]+$", "", label)
        if label in SUMMARY_MARKERS:
            return True
    return False


def looks_like_date(value: Any) -> bool:
    """Quickly test whether a candidate column contains date-like values."""
    if isinstance(value, (datetime, date)):
        return True
    if isinstance(value, (int, float)) and 20_000 <= value <= 80_000:
        return True
    text = clean_text(value)
    return bool(
        re.fullmatch(r"\d{4}[-/]\d{1,2}[-/]\d{1,2}", text)
        or re.fullmatch(r"\d{1,2}[-/]\d{1,2}(?:[-/]\d{1,2})?", text)
    )


def looks_like_data(value: Any) -> bool:
    """Distinguish ordinary data cells from blank/error cells."""
    if is_blank(value) or is_error(value):
        return False
    if isinstance(value, (int, float, datetime, date)):
        return True
    text = clean_text(value).replace(",", "").replace("%", "")
    try:
        float(text)
        return True
    except ValueError:
        return len(text) > 0


def coerce_excel_date(value: Any) -> datetime | date | None:
    """Return a real Excel date value instead of leaving date-looking text."""
    if isinstance(value, (datetime, date)):
        return value
    if isinstance(value, (int, float)) and 20_000 <= value <= 80_000:
        # openpyxl normally converts formatted serials already. Keep this path
        # defensive for vendor files whose date cells use General formatting.
        from openpyxl.utils.datetime import from_excel

        return from_excel(value)
    text = clean_text(value)
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass
    return None


# ---------------------------------------------------------------------------
# Header dictionary and matching
# ---------------------------------------------------------------------------

def read_dictionary(path: Path | None) -> dict[str, set[str]]:
    """Build target-column aliases from constants plus the supplied dictionary."""
    aliases = {target: {normalize(a) for a in values} for target, values in ALIASES.items()}
    if not path or not path.exists():
        return aliases

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["字典欄位說明"] if "字典欄位說明" in wb.sheetnames else wb.worksheets[0]
    canonical_to_target = {normalize(target): target for target in TARGET_COLUMNS}
    ignored_dictionary_values = {
        "na",
        "∅",
        "檔名",
        "品牌自填",
        "檔名[0]",
        "檔名[2]",
        "檔名[3]",
    }
    for row in ws.iter_rows(min_row=3, values_only=True):
        zh, en = row[1] if len(row) > 1 else None, row[2] if len(row) > 2 else None
        target = canonical_to_target.get(normalize(en))
        if not target:
            continue
        for value in (zh, en, *(row[7:12] if len(row) >= 12 else [])):
            alias = normalize(value)
            if not alias or alias in ignored_dictionary_values:
                continue
            claimed_elsewhere = any(
                alias in target_aliases
                for other_target, target_aliases in aliases.items()
                if other_target != target
            )
            if not claimed_elsewhere:
                aliases[target].add(alias)
    wb.close()
    return aliases


def match_header(header: Any, aliases: dict[str, set[str]]) -> str | None:
    """Map one source header to a system field, preferring exact matches."""
    key = normalize(header)
    if not key:
        return None
    exact = [target for target, values in aliases.items() if key in values]
    if exact:
        return exact[0]
    # Conservative fuzzy matching: only long aliases may be contained in a header.
    matches: list[tuple[int, str]] = []
    for target, values in aliases.items():
        for alias in values:
            if len(alias) >= 4 and (alias in key or key in alias):
                matches.append((len(alias), target))
    return max(matches)[1] if matches else None


# ---------------------------------------------------------------------------
# Ollama-assisted header detection
# ---------------------------------------------------------------------------

def compact_cell(value: Any, limit: int = 120) -> str | int | float | bool | None:
    """Make worksheet values small and JSON-safe before sending them to Ollama."""
    if value is None or isinstance(value, (int, float, bool)):
        return value
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return clean_text(value)[:limit]


def header_options_for_ollama(ws: Any, scan_rows: int) -> list[dict[str, Any]]:
    """Build a compact set of plausible rows plus data samples for one LLM call."""
    max_row = min(ws.max_row, scan_rows)
    max_col = min(ws.max_column, 100)
    base_aliases = {
        target: {normalize(alias) for alias in names}
        for target, names in ALIASES.items()
    }
    ranked: list[tuple[float, int, list[Any]]] = []
    for row_idx in range(1, max_row + 1):
        values = [ws.cell(row_idx, col).value for col in range(1, max_col + 1)]
        nonblank = [(col, value) for col, value in enumerate(values, 1) if not is_blank(value)]
        if len(nonblank) < 2:
            continue
        text_count = sum(isinstance(value, str) and not is_error(value) for _, value in nonblank)
        if text_count < 2:
            continue
        known_count = sum(match_header(value, base_aliases) is not None for _, value in nonblank)
        score = known_count * 8 + text_count / len(nonblank) * 5 + min(len(nonblank), 20) / 10
        ranked.append((score, row_idx, values))

    # Limit prompt size while retaining the strongest structurally plausible rows.
    selected = sorted(ranked, reverse=True)[:12]
    selected.sort(key=lambda item: item[1])
    options: list[dict[str, Any]] = []
    for _, row_idx, values in selected:
        headers = [
            {"column": col, "value": compact_cell(value)}
            for col, value in enumerate(values, 1)
            if not is_blank(value)
        ][:40]
        sample_rows: list[list[dict[str, Any]]] = []
        for sample_idx in range(row_idx + 1, min(ws.max_row, row_idx + 2) + 1):
            sample = [
                {"column": col, "value": compact_cell(ws.cell(sample_idx, col).value)}
                for col in range(1, max_col + 1)
                if not is_blank(ws.cell(sample_idx, col).value)
            ][:40]
            if sample:
                sample_rows.append(sample)
        options.append({"row": row_idx, "cells": headers, "sample_rows": sample_rows})
    return options


def call_ollama_json(prompt: str, config: OllamaConfig) -> dict[str, Any]:
    """Call Ollama's local chat API and return one validated JSON object."""
    schema = {
        "type": "object",
        "properties": {
            "header_row": {"type": "integer", "minimum": 0},
            "confidence": {"type": "number", "minimum": 0, "maximum": 1},
            "mappings": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "column": {"type": "integer", "minimum": 1},
                        "target": {"type": "string", "enum": list(TARGET_DESCRIPTIONS)},
                        "confidence": {"type": "number", "minimum": 0, "maximum": 1},
                    },
                    "required": ["column", "target", "confidence"],
                },
            },
        },
        "required": ["header_row", "confidence", "mappings"],
    }
    payload = {
        "model": config.model,
        "stream": False,
        "think": False,
        "format": schema,
        "options": {"temperature": 0},
        "messages": [
            {
                "role": "system",
                "content": OLLAMA_SYSTEM_PROMPT,
            },
            {"role": "user", "content": prompt},
        ],
    }
    request = urllib.request.Request(
        f"{config.url.rstrip('/')}/api/chat",
        data=json.dumps(payload, ensure_ascii=False).encode("utf-8"),
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(request, timeout=config.timeout) as response:
            result = json.loads(response.read().decode("utf-8"))
    except (OSError, TimeoutError, urllib.error.URLError, json.JSONDecodeError) as exc:
        raise RuntimeError(f"無法呼叫 Ollama：{exc}") from exc
    content = result.get("message", {}).get("content")
    if not isinstance(content, str):
        raise RuntimeError("Ollama 回應缺少 message.content")
    try:
        parsed = json.loads(content)
    except json.JSONDecodeError as exc:
        preview = clean_text(content)[:200]
        raise RuntimeError(f"Ollama 未回傳有效 JSON：{preview!r}") from exc
    if not isinstance(parsed, dict):
        raise RuntimeError("Ollama 回應不是 JSON object")
    return parsed


def llm_mapping_is_plausible(header: Any, target: str) -> bool:
    """Reject common LLM metric confusions even when confidence is overstated."""
    text = clean_text(header).lower()
    if target == "Video played to 25%":
        return "25" in text
    if target == "Video played to 50%":
        return "50" in text
    if target == "Video played to 75%":
        return "75" in text
    if target == "Video played to 100%":
        return any(marker in text for marker in ("100", "complete", "完整", "完成"))
    if target == "Campaign Type":
        return any(marker in text for marker in ("type", "format", "格式", "類型", "媒體 /", "媒體/"))
    return True


def mapping_is_sufficient(mapped: dict[int, str]) -> bool:
    """Require enough independent evidence to identify a real data table."""
    return len(set(mapped.values())) >= 2


def detect_candidate_with_ollama(
    ws: Any,
    aliases: dict[str, set[str]],
    scan_rows: int,
    config: OllamaConfig,
) -> Candidate | None:
    """Ask a local Ollama model to identify the header row and map its columns."""
    options = header_options_for_ollama(ws, scan_rows)
    if not options:
        return None
    prompt = render_ollama_user_prompt(
        target_descriptions=json.dumps(TARGET_DESCRIPTIONS, ensure_ascii=False, indent=2),
        options=json.dumps(options, ensure_ascii=False),
    )
    result = call_ollama_json(prompt, config)
    try:
        header_row = int(result.get("header_row", 0))
        confidence = max(0.0, min(float(result.get("confidence", 0)), 1.0))
    except (TypeError, ValueError) as exc:
        raise RuntimeError("Ollama 回傳的 header_row/confidence 格式錯誤") from exc
    if header_row < 1 or header_row > min(ws.max_row, scan_rows):
        return None

    max_col = min(ws.max_column, 100)
    raw_headers = [ws.cell(header_row, col).value for col in range(1, max_col + 1)]
    # Deterministic dictionary matches are more precise; Ollama fills only gaps.
    mapped: dict[int, str] = {}
    used_targets: set[str] = set()
    for col, value in enumerate(raw_headers, 1):
        if is_blank(value):
            continue
        target = match_header(value, aliases)
        if target and target not in used_targets:
            mapped[col] = target
            used_targets.add(target)

    mappings = result.get("mappings", [])
    if not isinstance(mappings, list):
        raise RuntimeError("Ollama 回傳的 mappings 格式錯誤")
    for item in mappings:
        if not isinstance(item, dict):
            continue
        try:
            col = int(item.get("column"))
        except (TypeError, ValueError):
            continue
        target = item.get("target")
        try:
            mapping_confidence = float(item.get("confidence", 0))
        except (TypeError, ValueError):
            continue
        if (
            1 <= col <= max_col
            and mapping_confidence >= 0.75
            and target in TARGET_DESCRIPTIONS
            and target not in used_targets
            and not is_blank(raw_headers[col - 1])
            and col not in mapped
            and llm_mapping_is_plausible(raw_headers[col - 1], target)
        ):
            mapped[col] = target
            used_targets.add(target)

    if not mapping_is_sufficient(mapped):
        return None

    checked = 0
    data_like = 0
    for row_idx in range(header_row + 1, min(ws.max_row, header_row + 8) + 1):
        filled = [
            ws.cell(row_idx, col).value
            for col in range(1, max_col + 1)
            if not is_blank(ws.cell(row_idx, col).value)
        ]
        if not filled:
            continue
        checked += 1
        if sum(looks_like_data(value) for value in filled) >= max(2, math.ceil(len(filled) * 0.5)):
            data_like += 1
    if checked == 0 or data_like / checked < 0.5:
        return None

    unmapped = [
        clean_text(value)
        for col, value in enumerate(raw_headers, 1)
        if not is_blank(value) and col not in mapped
    ]
    score = confidence * 100 + len(mapped) * 10 + data_like / checked * 10
    return Candidate(
        sheet=ws.title,
        header_row=header_row,
        score=round(score, 2),
        headers=[clean_text(value) for value in raw_headers],
        mapped=mapped,
        unmapped=unmapped,
        data_rows=0,
    )


# ---------------------------------------------------------------------------
# Table detection
# ---------------------------------------------------------------------------

def detect_candidates(ws: Any, aliases: dict[str, set[str]], scan_rows: int) -> list[Candidate]:
    """Find and score plausible header rows within one worksheet.

    A row must contain at least two independently mapped template fields and be
    followed by data-like rows. No single metric (such as Impressions) is
    required, so GA, TV, conversion-only and other valid reports are accepted.
    """
    candidates: list[Candidate] = []
    max_row = min(ws.max_row, scan_rows)
    max_col = min(ws.max_column, 100)
    for row_idx in range(1, max_row + 1):
        raw_headers = [ws.cell(row_idx, col).value for col in range(1, max_col + 1)]
        nonblank = [v for v in raw_headers if not is_blank(v)]
        if len(nonblank) < 2:
            continue

        mapped: dict[int, str] = {}
        used_targets: set[str] = set()
        unmapped: list[str] = []
        for col, value in enumerate(raw_headers, start=1):
            if is_blank(value):
                continue
            target = match_header(value, aliases)
            if target and target not in used_targets:
                mapped[col] = target
                used_targets.add(target)
            elif not target:
                unmapped.append(clean_text(value))

        if not mapping_is_sufficient(mapped):
            continue

        # Validate the next few rows. This rejects report titles such as
        # "保證曝光數 / 曝光達成率", which contain the keyword but are not tables.
        checked = 0
        data_like = 0
        date_like = 0
        for r in range(row_idx + 1, min(ws.max_row, row_idx + 8) + 1):
            values = [ws.cell(r, c).value for c in range(1, max_col + 1)]
            filled = [v for v in values if not is_blank(v)]
            if not filled:
                continue
            checked += 1
            if sum(looks_like_data(v) for v in filled) >= max(2, math.ceil(len(filled) * 0.5)):
                data_like += 1
            date_col = next((c for c, t in mapped.items() if t == "Date"), None)
            if date_col and looks_like_date(ws.cell(r, date_col).value):
                date_like += 1

        if checked == 0 or data_like / checked < 0.6:
            continue

        score = len(mapped) * 10 + (data_like / checked) * 10
        if "Date" in used_targets:
            score += 25 + min(date_like, 5) * 2
        if any(marker in ws.title.lower() for marker in PREFERRED_SHEETS):
            score += 30
        if row_idx > 1:
            score += 2  # Real vendor reports commonly have title bands above the table.
        candidates.append(
            Candidate(
                sheet=ws.title,
                header_row=row_idx,
                score=round(score, 2),
                headers=[clean_text(v) for v in raw_headers],
                mapped=mapped,
                unmapped=unmapped,
                data_rows=0,
            )
        )
    return candidates


def select_candidate(
    workbook: Any,
    aliases: dict[str, set[str]],
    scan_rows: int,
    sheet_name: str | None,
    ollama: OllamaConfig | None = None,
) -> Candidate | None:
    """Choose one table from one worksheet; never merge worksheets implicitly."""
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            available = ", ".join(workbook.sheetnames)
            raise ValueError(f"找不到工作表「{sheet_name}」。可用工作表：{available}")
        ws = workbook[sheet_name]
    else:
        ws = workbook.active

    if ollama and ollama.enabled:
        try:
            candidate = detect_candidate_with_ollama(ws, aliases, scan_rows, ollama)
            if candidate:
                return candidate
        except RuntimeError as exc:
            if not ollama.warned:
                print(f"[WARN] {exc}；改用原本的規則判別。", file=sys.stderr)
                ollama.warned = True
            ollama.enabled = False

    found = detect_candidates(ws, aliases, scan_rows)
    if not found:
        return None
    # Prefer the row that maps the most system fields. When a table repeats the
    # same header later in the sheet, start from the first occurrence.
    return max(found, key=lambda item: (len(item.mapped), -item.header_row, item.score))


def infer_context(ws: Any, candidate: Candidate) -> tuple[str | None, str | None]:
    """Read a campaign hint explicitly labelled above the detected table."""
    campaign = None
    for r in range(1, candidate.header_row):
        row = [clean_text(ws.cell(r, c).value) for c in range(1, min(ws.max_column, 30) + 1)]
        for c, text in enumerate(row):
            low = text.lower()
            if "客戶案名" in text or "campaign name" in low:
                campaign = next((v for v in row[c + 1 :] if v), campaign)
            if "廣告素材" in text and text not in {"廣告素材"}:
                campaign = re.sub(r"^.*?廣告素材[_：: ]*", "", text).strip("_") or campaign
    return campaign, None


# ---------------------------------------------------------------------------
# Row cleaning and workbook output
# ---------------------------------------------------------------------------

def iter_clean_rows(ws: Any, candidate: Candidate) -> Iterable[dict[str, Any]]:
    """Yield standardized records and discard non-data/trailing rows."""
    campaign_hint, ad_type_hint = infer_context(ws, candidate)
    blank_streak = 0
    for r in range(candidate.header_row + 1, ws.max_row + 1):
        values = [ws.cell(r, c).value for c in range(1, min(ws.max_column, 100) + 1)]
        if all(is_blank(v) for v in values):
            blank_streak += 1
            if blank_streak >= 5:
                break
            continue
        blank_streak = 0
        if is_summary_row(values):
            continue

        record = {column: None for column in TARGET_COLUMNS}
        record["Campaign name"] = campaign_hint
        record["Campaign Type"] = ad_type_hint
        for col, target in candidate.mapped.items():
            value = ws.cell(r, col).value
            record[target] = None if is_error(value) or is_blank(value) else value

        if record["Date"] is not None:
            record["Date"] = coerce_excel_date(record["Date"])

        mapped_values = [record[target] for target in candidate.mapped.values()]
        if not any(not is_blank(value) and looks_like_data(value) for value in mapped_values):
            continue
        if "Date" in candidate.mapped.values() and record["Date"] is None:
            continue
        yield record


def populated_columns(records: Iterable[dict[str, Any]]) -> list[str]:
    """Return non-empty output columns in canonical template order."""
    rows = list(records)
    return [
        column
        for column in TARGET_COLUMNS
        if any(not is_blank(record.get(column)) for record in rows)
    ]


def format_output(
    ws: Any,
    row_count: int,
    output_columns: Iterable[str] | None = None,
) -> None:
    """Apply a small, consistent format to the cleaned output worksheet."""
    columns = list(output_columns or (cell.value for cell in ws[1] if cell.value))
    if not columns:
        return
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    last_column = get_column_letter(len(columns))
    ws.auto_filter.ref = f"A1:{last_column}{max(row_count + 1, 1)}"

    widths = {
        "Date": 13,
        "Campaign name": 30,
        "Adset name": 30,
        "Ad Free Form": 30,
        "Final URL": 40,
        "Impressions": 14,
        "Clicks (all)": 14,
        "Spent (TWD)": 16,
        "Campaign Type": 18,
    }
    column_letters = {
        target: get_column_letter(index)
        for index, target in enumerate(columns, start=1)
    }
    for target, column_letter in column_letters.items():
        ws.column_dimensions[column_letter].width = widths.get(target, 18)

    if "Date" in column_letters:
        for cell in ws[column_letters["Date"]][1:]:
            cell.number_format = "yyyy-mm-dd"
    for target in ("Bounce Rate", "TVR", "10 Second TVR"):
        if target in column_letters:
            for cell in ws[column_letters[target]][1:]:
                cell.number_format = "0.00%"
    for target in (
        "Reach",
        "Impressions",
        "Clicks (all)",
        "Link clicks (Web Clicks)",
        "Views",
        '3" Video Views',
        '15" Video Views (ThruPlays)',
        "TrueView: Views",
        "Video played to 25%",
        "Video played to 50%",
        "Video played to 75%",
        "Video played to 100%",
    ):
        if target in column_letters:
            for cell in ws[column_letters[target]][1:]:
                cell.number_format = "#,##0"


def _format_audit_sheet(
    ws: Any,
    row_count: int,
    widths: dict[str, int],
) -> None:
    """Apply compact, consistent formatting to an audit worksheet."""
    header_fill = PatternFill("solid", fgColor="5B6573")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    last_column = get_column_letter(ws.max_column)
    ws.auto_filter.ref = f"A1:{last_column}{max(row_count + 1, 1)}"

    header_columns = {
        cell.value: get_column_letter(index)
        for index, cell in enumerate(ws[1], start=1)
    }
    for header, width in widths.items():
        column = header_columns.get(header)
        if column:
            ws.column_dimensions[column].width = width


def consolidate_cleaned_workbooks(
    cleaned_paths: Iterable[Path],
    records: Iterable[AuditRecord],
    output_path: Path,
) -> int:
    """Combine cleaned data and audit details into one Excel workbook."""
    audit_records = list(records)
    combined_records: list[dict[str, Any]] = []

    for cleaned_path in cleaned_paths:
        cleaned = load_workbook(cleaned_path, data_only=True, read_only=True)
        try:
            source_ws = (
                cleaned["cleaned_data"]
                if "cleaned_data" in cleaned.sheetnames
                else cleaned.worksheets[0]
            )
            rows = source_ws.iter_rows(values_only=True)
            headers = [clean_text(value) for value in next(rows, ())]
            invalid_headers = [
                header for header in headers if header and header not in TARGET_COLUMNS
            ]
            if invalid_headers:
                raise ValueError(
                    f"{cleaned_path.name} 含有非 template 欄位："
                    f"{'、'.join(invalid_headers)}"
                )
            for row in rows:
                record = {
                    header: value
                    for header, value in zip(headers, row)
                    if header in TARGET_COLUMNS and not is_blank(value)
                }
                if record:
                    combined_records.append(record)
        finally:
            cleaned.close()

    output_columns = populated_columns(combined_records)
    workbook = Workbook()
    data_ws = workbook.active
    data_ws.title = "cleaned_data"
    data_ws.append(output_columns)
    for record in combined_records:
        data_ws.append([record.get(column) for column in output_columns])
    total_rows = len(combined_records)
    format_output(data_ws, total_rows, output_columns)

    audit_ws = workbook.create_sheet("cleaning_audit")
    audit_ws.append(
        [
            "input_file",
            "sheet",
            "header_row",
            "status",
            "score",
            "data_rows",
            "mapped_columns",
            "unmapped_columns",
            "output_file",
        ]
    )
    for record in audit_records:
        audit_ws.append(
            [
                record.input_file,
                record.sheet,
                record.header_row,
                record.status,
                record.score,
                record.data_rows,
                json.dumps(record.mapped_columns, ensure_ascii=False),
                "、".join(record.unmapped_columns),
                record.output_file,
            ]
        )
    _format_audit_sheet(
        audit_ws,
        len(audit_records),
        {
            "input_file": 28,
            "sheet": 20,
            "header_row": 12,
            "status": 28,
            "score": 12,
            "data_rows": 12,
            "mapped_columns": 60,
            "unmapped_columns": 36,
            "output_file": 28,
        },
    )
    for column in ("G", "H"):
        for cell in audit_ws[column][1:]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    unmapped_ws = workbook.create_sheet("unmapped_columns")
    unmapped_ws.append(
        ["input_file", "sheet", "header_row", "status", "unmapped_column"]
    )
    unmapped_rows = 0
    for record in audit_records:
        if record.unmapped_columns:
            for column in record.unmapped_columns:
                unmapped_ws.append(
                    [
                        record.input_file,
                        record.sheet,
                        record.header_row,
                        record.status,
                        column,
                    ]
                )
                unmapped_rows += 1
        elif record.status != "成功":
            unmapped_ws.append(
                [
                    record.input_file,
                    record.sheet,
                    record.header_row,
                    record.status,
                    "",
                ]
            )
            unmapped_rows += 1
    _format_audit_sheet(
        unmapped_ws,
        unmapped_rows,
        {
            "input_file": 28,
            "sheet": 20,
            "header_row": 12,
            "status": 28,
            "unmapped_column": 36,
        },
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    workbook.close()
    return total_rows


def list_workbook_sheets(workbook_data: bytes) -> list[str]:
    """Return worksheet names in workbook order without changing the file."""
    workbook = load_workbook(io.BytesIO(workbook_data), data_only=True, read_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def clean_workbook_sheets(
    input_path: Path,
    output_path: Path,
    aliases: dict[str, set[str]],
    scan_rows: int,
    sheet_names: Iterable[str],
    ollama: OllamaConfig | None = None,
) -> tuple[list[AuditRecord], int]:
    """Clean explicitly selected worksheets into one standardized output table.

    Each worksheet is detected and audited independently. Originating file and
    worksheet remain traceable through the audit output.
    """
    selected_sheets = list(dict.fromkeys(name for name in sheet_names if name))
    if not selected_sheets:
        raise ValueError("請至少選擇一個工作表。")

    wb_values = load_workbook(input_path, data_only=True, read_only=False)
    missing_sheets = [
        sheet_name
        for sheet_name in selected_sheets
        if sheet_name not in wb_values.sheetnames
    ]
    if missing_sheets:
        available = "、".join(wb_values.sheetnames)
        wb_values.close()
        missing = "、".join(missing_sheets)
        raise ValueError(f"找不到工作表「{missing}」。可用工作表：{available}")

    output = Workbook()
    out_ws = output.active
    out_ws.title = "cleaned_data"
    audit: list[AuditRecord] = []
    cleaned_records: list[dict[str, Any]] = []
    total_rows = 0

    try:
        for sheet_name in selected_sheets:
            try:
                candidate = select_candidate(
                    wb_values,
                    aliases,
                    scan_rows,
                    sheet_name,
                    ollama,
                )
                if not candidate:
                    audit.append(
                        AuditRecord(
                            str(input_path),
                            sheet_name,
                            None,
                            "找不到資料表表頭",
                            None,
                            0,
                            {},
                            [],
                            None,
                        )
                    )
                    continue

                ws = wb_values[candidate.sheet]
                rows = list(iter_clean_rows(ws, candidate))
                cleaned_records.extend(rows)
                total_rows += len(rows)
                candidate.data_rows = len(rows)
                mapped_columns = {
                    candidate.headers[col - 1]: target
                    for col, target in candidate.mapped.items()
                }
                audit.append(
                    AuditRecord(
                        str(input_path),
                        candidate.sheet,
                        candidate.header_row,
                        "成功" if rows else "表頭找到但沒有有效資料",
                        candidate.score,
                        len(rows),
                        mapped_columns,
                        candidate.unmapped,
                        str(output_path) if rows else None,
                    )
                )
            except Exception as exc:
                audit.append(
                    AuditRecord(
                        str(input_path),
                        sheet_name,
                        None,
                        f"處理失敗：{exc}",
                        None,
                        0,
                        {},
                        [],
                        None,
                    )
                )
    finally:
        wb_values.close()

    if total_rows:
        output_columns = populated_columns(cleaned_records)
        out_ws.append(output_columns)
        for record in cleaned_records:
            out_ws.append([record[column] for column in output_columns])
        format_output(out_ws, total_rows, output_columns)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output.save(output_path)
    output.close()
    return audit, total_rows


def clean_workbook(
    input_path: Path,
    output_path: Path,
    aliases: dict[str, set[str]],
    scan_rows: int,
    sheet_name: str | None,
    ollama: OllamaConfig | None = None,
) -> tuple[list[AuditRecord], int]:
    """Clean one workbook/worksheet and save its standardized output file."""
    selected_sheet = sheet_name
    if not selected_sheet:
        workbook = load_workbook(input_path, data_only=True, read_only=True)
        try:
            selected_sheet = workbook.active.title
        finally:
            workbook.close()
    return clean_workbook_sheets(
        input_path=input_path,
        output_path=output_path,
        aliases=aliases,
        scan_rows=scan_rows,
        sheet_names=[selected_sheet],
        ollama=ollama,
    )


def discover_inputs(path: Path) -> list[Path]:
    """Return one input workbook or eligible workbooks inside a folder."""
    if path.is_file():
        return [path]
    return sorted(
        p for p in path.glob("*.xlsx")
        if not p.name.startswith("~$") and "template" not in p.stem.lower() and "cleaned" not in p.stem.lower()
    )


def write_audit(records: list[AuditRecord], output_dir: Path) -> None:
    """Write detailed JSON audit data and a simple unmapped-column CSV."""
    output_dir.mkdir(parents=True, exist_ok=True)
    json_path = output_dir / "cleaning_audit.json"
    csv_path = output_dir / "unmapped_columns.csv"
    json_path.write_text(json.dumps([asdict(r) for r in records], ensure_ascii=False, indent=2), encoding="utf-8")
    with csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["input_file", "sheet", "header_row", "status", "unmapped_column"])
        for record in records:
            if record.unmapped_columns:
                for column in record.unmapped_columns:
                    writer.writerow([record.input_file, record.sheet, record.header_row, record.status, column])
            elif record.status != "成功":
                writer.writerow([record.input_file, record.sheet, record.header_row, record.status, ""])


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    """Define and parse command-line options."""
    parser = argparse.ArgumentParser(description="偵測未知 Excel 報表中的資料表並統一媒體欄位")
    parser.add_argument("input", type=Path, help="單一 .xlsx 檔或包含 .xlsx 的資料夾")
    parser.add_argument("-o", "--output-dir", type=Path, default=Path("cleaned_output"), help="輸出資料夾")
    parser.add_argument(
        "--dictionary", type=Path, default=Path("Report Template & All Format 字典.xlsx"), help="欄位字典 Excel",
    )
    parser.add_argument("--scan-rows", type=int, default=100, help="每張工作表最多掃描幾列尋找表頭")
    parser.add_argument("--sheet", help="要清理的工作表名稱；未指定時使用 Excel 目前選取的工作表")
    parser.add_argument("--ollama-model", default="qwen3.5:9b", help="本機 Ollama 模型（預設：qwen3.5:9b）")
    parser.add_argument("--ollama-url", default="http://127.0.0.1:11434", help="Ollama API 網址")
    parser.add_argument("--ollama-timeout", type=float, default=120, help="等待 Ollama 回應的秒數")
    parser.add_argument("--no-ollama", action="store_true", help="停用 Ollama，僅使用原本規則判別")
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    """CLI entry point: discover inputs, clean them, and report results."""
    args = parse_args(argv)
    inputs = discover_inputs(args.input)
    if not inputs:
        print(f"找不到可處理的 .xlsx：{args.input}", file=sys.stderr)
        return 2
    aliases = read_dictionary(args.dictionary)
    ollama = OllamaConfig(
        model=args.ollama_model,
        url=args.ollama_url,
        timeout=args.ollama_timeout,
        enabled=not args.no_ollama,
    )
    all_audit: list[AuditRecord] = []
    succeeded = 0
    for input_path in inputs:
        output_path = args.output_dir / f"{input_path.stem}_cleaned.xlsx"
        try:
            audit, rows = clean_workbook(
                input_path, output_path, aliases, args.scan_rows, args.sheet, ollama
            )
        except ValueError as exc:
            print(f"[ERROR] {input_path.name}: {exc}", file=sys.stderr)
            continue
        all_audit.extend(audit)
        if rows:
            succeeded += 1
            print(f"[OK] {input_path.name}: {rows} rows -> {output_path}")
        else:
            print(f"[WARN] {input_path.name}: 沒有輸出有效資料", file=sys.stderr)
    write_audit(all_audit, args.output_dir)
    print(f"完成：{succeeded}/{len(inputs)} 個檔案；稽核報告位於 {args.output_dir}")
    return 0 if succeeded else 1


if __name__ == "__main__":
    raise SystemExit(main())
